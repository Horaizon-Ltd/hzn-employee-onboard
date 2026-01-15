import json
import base64
import pandas as pd
import numpy as np
import pytesseract
import os
import tempfile
import traceback
import re
import logging
import gc
import boto3

from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from botocore.config import Config

from pdf2image import convert_from_path
from pathlib import Path
from typing import Dict
from pytesseract import Output
from openpyxl import load_workbook
from io import BytesIO

# Configure logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Pytesseract Configs
PYTESSERACT_CONFIG = "--oem 3 --psm 4"
PYTESSERACT_LANG = "dan"

# Email coordinate for employee general PDF
general_employee_email_coordinate = (1, 1107, 1651, 1189)

def process_single_file(file_path, file_type, file_format, file_name):
    try:
        logger.info(f"Processing {file_type} ({file_format}): {file_name}")
        result_df = process_file_by_type(file_path, file_type, file_format)
        logger.info(f"Completed {file_type}: {len(result_df)} records, {len(result_df.columns)} columns")

        gc.collect()

        return (file_type, result_df, None)
    except Exception as e:
        logger.error(f"Error processing {file_type}: {e}")
        logger.error(traceback.format_exc())
        gc.collect()
        return (file_type, pd.DataFrame({'error': [str(e)]}), str(e))
    
def parse_european_float_to_numeric(value):
    """
    Convert European-formatted number string to Python float.
    E.g., "1.234,56" -> 1234.56, "12,5" -> 12.5
    Returns NaN for invalid values (will be cleaned up later in formatting)
    """
    if pd.isna(value) or value == "" or value is None:
        return np.nan

    if isinstance(value, str):
        value_clean = value.replace("%", "").strip()

        # Check for NaN-like strings
        if value_clean.lower() in ['nan', 'none', '<na>', 'nat', '']:
            return np.nan

        try:
            # European format: dots are thousands separators, comma is decimal
            cleaned = value_clean.replace(".", "").replace(",", ".")
            num = float(cleaned)
            return num
        except (ValueError, TypeError):
            try:
                num = float(value_clean)
                return num
            except (ValueError, TypeError):
                return np.nan
    else:
        try:
            return float(value)
        except (ValueError, TypeError):
            return np.nan

def clean_float_string(value):
    """
    Clean up float strings to ensure proper European format:
    - Remove thousand separators (.)
    - Ensure only one comma for decimals (fix OCR issues)
    - Ensure exactly 2 decimal places
    - Remove NaN string values
    """
    if pd.isna(value) or value == "" or value is None:
        return ''

    value_str = str(value).strip()

    # Skip if empty or NaN-like string - return empty string
    if not value_str or value_str.lower() in ['', 'nan', 'none', '<na>', 'nat']:
        return ''

    # Remove thousand separators (dots) - only keep commas
    # Remove all dots (thousand separators)
    value_str = value_str.replace('.', '')

    # Fix double commas (OCR errors) - keep only the first comma
    if value_str.count(',') > 1:
        parts = value_str.split(',')
        # Keep first part and first decimal part only
        value_str = parts[0] + ',' + ''.join(parts[1:])

    # Ensure exactly 2 decimal places (only for numeric strings)
    # Check if it looks like a number (contains only digits, comma, and minus)
    if all(c.isdigit() or c in [',', '-'] for c in value_str):
        if ',' in value_str:
            parts = value_str.split(',')
            integer_part = parts[0]
            decimal_part = parts[1] if len(parts) > 1 else ''

            # Trim or pad decimal part to exactly 2 digits
            if len(decimal_part) == 0:
                decimal_part = '00'
            elif len(decimal_part) == 1:
                decimal_part = decimal_part + '0'
            elif len(decimal_part) > 2:
                decimal_part = decimal_part[:2]

            value_str = f"{integer_part},{decimal_part}"
        else:
            # No comma found - check if it's a whole number
            if value_str.lstrip('-').isdigit():
                value_str = f"{value_str},00"

    return value_str

def format_european_float(value, force_decimal=True):
    if pd.isna(value) or value == "" or value is None:
        return ''

    if isinstance(value, str):
        value_clean = value.replace("%", "").strip()

        # Check if it's a NaN-like string
        if value_clean.lower() in ['nan', 'none', '<na>', 'nat', '']:
            return ''

        try:
            cleaned = value_clean.replace(".", "").replace(",", ".")
            num = float(cleaned)
        except (ValueError, TypeError):
            try:
                num = float(value_clean)
            except (ValueError, TypeError):
                return value
    else:
        try:
            num = float(value)
        except (ValueError, TypeError):
            return value

    # ALWAYS format with exactly 2 decimal places
    formatted = f"{num:.2f}"

    # Replace decimal point with comma
    formatted = formatted.replace(".", ",")

    return formatted

def lambda_handler(event, context):
    try:
        logger.info("=== Lambda Invocation Started ===")
        logger.info(f"Event keys: {list(event.keys())}")

        # Parse request body
        if 'body' in event:
            logger.info("Parsing body from event")
            if event.get('isBase64Encoded', False):
                body = base64.b64decode(event['body']).decode('utf-8')
            else:
                body = event['body']
            request_data = json.loads(body)
        else:
            logger.info("Using event directly as request_data")
            request_data = event

        # Check if files are provided via S3 keys or base64
        s3_files = request_data.get('s3_files', [])
        base64_files = request_data.get('files', [])

        # Create temp directory
        temp_dir = tempfile.mkdtemp()
        processed_data = {}
        file_tasks = []

        # Handle S3-based files (new approach)
        if s3_files:
            logger.info(f"Processing {len(s3_files)} files from S3")
            uploads_bucket = os.environ.get('UPLOADS_BUCKET')

            if not uploads_bucket:
                raise ValueError("UPLOADS_BUCKET environment variable not set")

            s3_client = boto3.client('s3', config=Config(signature_version='s3v4'))

            for s3_file in s3_files:
                s3_key = s3_file.get('s3_key')
                file_type = s3_file.get('file_type')

                if not s3_key or not file_type:
                    logger.warning(f"Skipping incomplete S3 file info: {s3_file}")
                    continue

                # Extract filename and format from S3 key
                file_name = os.path.basename(s3_key)
                file_format = file_name.split('.')[-1].lower()

                logger.info(f"Downloading {file_type} from S3: {s3_key}")

                # Download file from S3
                input_file = os.path.join(temp_dir, file_name)
                try:
                    s3_client.download_file(uploads_bucket, s3_key, input_file)
                    logger.info(f"Downloaded {file_name} successfully")

                    file_tasks.append((input_file, file_type, file_format, file_name))
                except Exception as e:
                    logger.error(f"Failed to download {s3_key}: {e}")
                    continue

        # Handle base64 files (backward compatibility)
        elif base64_files:
            logger.info(f"Processing {len(base64_files)} files from base64")

            for file_info in base64_files:
                file_type = file_info.get('file_type')
                file_name = file_info.get('file_name')
                file_format = file_info.get('file_format')
                file_data = file_info.get('file_data')

                if not all([file_type, file_name, file_format, file_data]):
                    logger.warning(f"Skipping incomplete file info: {file_info}")
                    continue

                # Decode and save file
                decoded_data = base64.b64decode(file_data)
                input_file = os.path.join(temp_dir, file_name)

                with open(input_file, 'wb') as f:
                    f.write(decoded_data)

                file_tasks.append((input_file, file_type, file_format, file_name))
                logger.info(f"Prepared {file_type}: {file_name}")
        else:
            raise ValueError("Either s3_files or files array is required")

        if not file_tasks:
            raise ValueError("No valid files to process")

        logger.info(f"Total files prepared for processing: {len(file_tasks)}")
        logger.info(f"Starting smart parallel processing of {len(file_tasks)} files...")

        # Separate low-memory (Excel) and high-memory (PDF) tasks
        excel_tasks = [(f, t, fmt, n) for f, t, fmt, n in file_tasks if fmt == 'xlsx']
        pdf_tasks = [(f, t, fmt, n) for f, t, fmt, n in file_tasks if fmt == 'pdf']

        logger.info(f"Excel files: {len(excel_tasks)}, PDF files: {len(pdf_tasks)}")

        # Process Excel files in parallel (low memory usage)
        if excel_tasks:
            logger.info("Processing Excel files in parallel...")
            with ThreadPoolExecutor(max_workers=len(excel_tasks)) as executor:
                future_to_file = {
                    executor.submit(process_single_file, *task): task
                    for task in excel_tasks
                }

                for future in as_completed(future_to_file):
                    task = future_to_file[future]
                    try:
                        file_type, result_df, error = future.result()
                        processed_data[file_type] = result_df
                        if error:
                            logger.error(f"Error in {file_type}: {error}")
                        else:
                            logger.info(f"✓ Completed {file_type}: {len(result_df)} records")
                    except Exception as e:
                        logger.error(f"Exception processing {task[3]}: {e}")
                        logger.error(traceback.format_exc())

        if pdf_tasks:
            pdf_workers = 1
            logger.info(f"Processing {len(pdf_tasks)} PDF files sequentially for quality...")
            with ThreadPoolExecutor(max_workers=pdf_workers) as executor:
                future_to_file = {
                    executor.submit(process_single_file, *task): task
                    for task in pdf_tasks
                }

                for future in as_completed(future_to_file):
                    task = future_to_file[future]
                    try:
                        file_type, result_df, error = future.result()
                        processed_data[file_type] = result_df
                        if error:
                            logger.error(f"Error in {file_type}: {error}")
                        else:
                            logger.info(f"✓ Completed {file_type}: {len(result_df)} records")
                    except Exception as e:
                        logger.error(f"Exception processing {task[3]}: {e}")
                        logger.error(traceback.format_exc())

        for input_file, _, _, _ in file_tasks:
            try:
                os.remove(input_file)
            except Exception as e:
                logger.warning(f"Could not remove {input_file}: {e}")

        logger.info(f"Processed data summary: {list(processed_data.keys())}")

        logger.info("Starting merge and mapping...")
        final_result = merge_and_map_data(processed_data)

        if final_result is not None and not final_result.empty:
            logger.info(f"Final result: {len(final_result)} rows, {len(final_result.columns)} columns")

            # Debug: Check CPR values right after mapping
            for col in final_result.columns:
                if col and 'CPR' in str(col).upper():
                    sample_cpr = final_result[col].dropna().head(10).astype(str).tolist()
                    logger.info(f"🔍 DEBUG: After mapping, column '{col}' dtype={final_result[col].dtype}, values={sample_cpr}")
        else:
            logger.error("Merge and map returned empty result")
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({
                    'success': False,
                    'error': 'Missing key data for processing.'
                })
            }

        # Generate Excel output (better preserves data types and leading zeros)
        logger.info("Generating Excel output...")

        # Create Excel file in memory using BytesIO
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        excel_buffer = BytesIO()

        # Write initial Excel file
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            final_result.to_excel(writer, index=False, sheet_name='Sheet1')

        # Reopen to set text format for columns with leading zeros
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Check first few rows to identify columns with leading zeros
        logger.info("Setting text format for columns with leading zeros...")
        logger.info(f"Total columns to check: {len(final_result.columns)}")

        # Known ID column keywords that should always be formatted as text
        id_keywords = ['CPR', 'NUMMER', 'MEDARBEJDER', 'POSTNUMMER', 'POST', 'CVR', 'KONTO']

        for col_idx, col_name in enumerate(final_result.columns, start=1):
            # Skip if column name is None or empty
            if col_name is None or col_name == '':
                logger.warning(f"Skipping column at index {col_idx} - column name is None or empty")
                continue

            col_name_upper = str(col_name).upper()

            # Sample first 50 non-empty values
            sample_values = final_result[col_name].dropna().head(50).astype(str)

            # Log sample values for CPR-related columns
            if 'CPR' in col_name_upper or 'NUMMER' in col_name_upper:
                logger.info(f"Column '{col_name}' dtype: {final_result[col_name].dtype}, sample values (first 10): {list(sample_values.head(10))}")

            # Check if any values have leading zeros
            has_leading_zeros = any(
                str(v).startswith('0') and len(str(v)) > 1 and str(v)[1:].replace('.','').replace(',','').isdigit()
                for v in sample_values if str(v) not in ['', 'nan', 'None']
            )

            # Check if column name contains ID keywords (force text format)
            is_id_column = any(keyword in col_name_upper for keyword in id_keywords)

            # Check if column contains European-formatted numbers (e.g., "123,56" or "1234,5")
            # These should be kept as text to preserve formatting
            has_european_format = any(
                ',' in str(v) and str(v).replace(',','').replace('-','').isdigit()
                for v in sample_values if str(v) not in ['', 'nan', 'None']
            )

            # Additional logging for debugging
            if has_european_format:
                logger.info(f"European format detected in '{col_name}': {list(sample_values.head(3))}")

            if has_leading_zeros or is_id_column or has_european_format:
                col_letter = get_column_letter(col_idx)
                if has_leading_zeros:
                    reason = "detected leading zeros"
                elif is_id_column:
                    reason = "ID column keyword match"
                else:
                    reason = "European number format detected (preserving thousand separators and comma decimals)"
                logger.info(f"Applying text format to column '{col_name}' ({col_letter}) - {reason}")
                # Apply text format to all cells in this column (skip header)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    cell.number_format = '@'  # Text format

        # Save back to buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_output = excel_buffer.getvalue()

        logger.info(f"Excel generated with text formatting, size: {len(excel_output)} bytes")

        logger.info("Uploading Excel to S3...")

        s3_client = boto3.client('s3', config=Config(signature_version='s3v4'))
        bucket_name = os.environ.get('RESULTS_BUCKET')

        if not bucket_name:
            raise ValueError("RESULTS_BUCKET environment variable not set")

        # Generate unique filename with timestamp
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        s3_key = f"results/{timestamp}_danlon_processed_output.xlsx"

        # Upload Excel to S3
        s3_client.put_object(
            Bucket=bucket_name,
            Key=s3_key,
            Body=excel_output,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ContentDisposition='attachment; filename="danlon_processed_output.xlsx"'
        )
        logger.info(f"Excel uploaded to S3: s3://{bucket_name}/{s3_key}")

        download_url = s3_client.generate_presigned_url(
            'get_object',
            Params={
                'Bucket': bucket_name,
                'Key': s3_key,
                'ResponseContentDisposition': 'attachment; filename="danlon_processed_output.xlsx"',
                'ResponseContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            },
            ExpiresIn=3600  # 1 hour
        )
        logger.info("Generated presigned download URL")

        # Cleanup temp directory
        try:
            os.rmdir(temp_dir)
            logger.info("Temp directory cleaned up")
        except Exception as e:
            logger.warning(f"Could not remove temp dir: {e}")
        response_body = {
            'success': True,
            'download_url': download_url,
            'file_name': "danlon_processed_output.xlsx",
            'records_processed': len(final_result),
            'files_processed': len(file_tasks),
            'processing_summary': {k: len(v) for k, v in processed_data.items()},
            's3_key': s3_key
        }

        logger.info(f"=== Processing Complete ===")
        logger.info(f"Records processed: {len(final_result)}")
        logger.info(f"Files processed: {len(file_tasks)}")
        logger.info(f"Summary: {response_body['processing_summary']}")

        # Create JSON response
        logger.info("Converting response to JSON...")
        response_json = json.dumps(response_body)
        logger.info(f"Response size: {len(response_json)} bytes")

        final_response = {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json'
            },
            'body': response_json
        }
        return final_response

    except Exception as e:
        logger.error(f"=== Error Occurred ===")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error: {str(e)}")
        logger.error(traceback.format_exc())

        error_response = {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
            },
            'body': json.dumps({
                'success': False,
                'error': str(e),
                'error_type': type(e).__name__
            })
        }
        return error_response
    
def extract_table_from_section(
    image,
    columns_dict: dict | list,
    y_threshold: int = 10,
    lang: str = "dan",
    optional: bool = False
) -> pd.DataFrame | None:
    data = pytesseract.image_to_data(image, lang=PYTESSERACT_LANG, output_type=Output.DATAFRAME, config=PYTESSERACT_CONFIG)

    if optional:
        lines = data["text"].dropna().tolist()

        if any(any(col in line for col in columns_dict.keys()) for line in lines):
            print("Columns found for optional section")
        else:
            print("Column not found for optional section")
            return None

    data = data[data['conf'] > 0]
    data = data.dropna(subset=['text'])
    data["x_center"] = data["left"] + data["width"] / 2
    data["y_center"] = data["top"] + data["height"] / 2

    if type(columns_dict) == list:
        if "Saldo" in data['text'].values:
            columns_dict = columns_dict[0]
        else:
            columns_dict = columns_dict[1]

    # Assign each word to a column based on its horizontal x position
    def assign_column(x):
        for col, (xmin, xmax) in columns_dict.items():
            if xmin <= x <= xmax:
                return col
        return None
    
    data["column"] = data["x_center"].apply(assign_column)
    data = data.sort_values(by="y_center").reset_index(drop=True)

    rows = []
    current_row_y = None
    current_row = {col: [] for col in columns_dict}

    for _, w in data.iterrows():
        if w["text"] in columns_dict.keys():  # skip header row
            continue

        if current_row_y is None:
            current_row_y = w["y_center"]

        # Detect new row
        if abs(w["y_center"] - current_row_y) > y_threshold:
            # Merge words within each column (sorted left-to-right)
            sorted_row = {}
            for col, words in current_row.items():
                words_sorted = sorted(words, key=lambda x: x["left"])
                sorted_row[col] = " ".join([x["text"] for x in words_sorted])
            rows.append(sorted_row)

            current_row = {col: [] for col in columns_dict}
            current_row_y = w["y_center"]

        col = w["column"]
        if col:
            current_row[col].append(w)

    # Append last row
    sorted_row = {}
    for col, words in current_row.items():
        words_sorted = sorted(words, key=lambda x: x["left"])
        sorted_row[col] = " ".join([x["text"] for x in words_sorted])
    rows.append(sorted_row)

    df = pd.DataFrame(rows)
    df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)
    df = df.dropna(how="all").reset_index(drop=True)

    # drop row if first col no val
    first_col = df.columns[0]
    df = df[df[first_col].astype(bool)]
    df = df.reset_index(drop=True)

    return df

def process_file_by_type(file_path, file_type, file_format):
    """
    Route to appropriate processor based on file type
    """
    if file_type == "employee_payslip" and file_format == "pdf":
        return process_payslip(file_path, file_format)
    elif file_type == "employee_active" and file_format in ["xlsx", "xls"]:
        # Read all columns as strings to preserve leading zeros in CPR, postal codes, etc.
        df = pd.read_excel(file_path, dtype=str)
        logger.info(f"Read employee_active with {len(df)} rows, all columns as strings")
        return df
    elif file_type == "employee_holiday" and file_format in ["xlsx", "xls"]:
        # Read all columns as strings to preserve leading zeros
        df = pd.read_excel(file_path, dtype=str)
        logger.info(f"Read employee_holiday with {len(df)} rows, all columns as strings")

        # Convert American number format to European format by swapping , and .
        # American: 29,293.55 or 0.00 (period = decimal separator)
        # European: 29.293,55 or 0,00 (comma = decimal separator)
        for col in df.columns:
            # Check if column contains number-like values with period (American decimal separator)
            sample_values = df[col].dropna().head(20)
            has_american_format = any(
                ('.' in str(v) or ',' in str(v)) and str(v).replace(',', '').replace('.', '').replace('-', '').replace(' ', '').isdigit()
                for v in sample_values if str(v) not in ['', 'nan', 'None']
            )

            if has_american_format:
                # Simply swap , and . for the entire column
                # "29,293.55" -> "29.293,55"
                # "0.00" -> "0,00"
                df[col] = df[col].apply(lambda x:
                    str(x).replace(',', 'TEMP').replace('.', ',').replace('TEMP', '.')
                    if pd.notna(x) and (('.' in str(x)) or (',' in str(x)))
                    else x
                )
                print(f"Converted column '{col}' from American to European number format")
                logger.info(f"Converted column '{col}' from American to European number format")

        return df
    elif file_type == "employee_general" and file_format == "pdf":
        return process_employee_general(file_path, file_format)
    elif file_type == "employee_list" and file_format == "pdf":
        return process_employee_list(file_path, file_format)
    else:
        raise ValueError(f"Unsupported file type: {file_type} with format: {file_format}")
    
def extract_column_from_df(df: pd.DataFrame) -> Dict:
    result = {}

    if df.empty or df.shape[1] < 2:
        return result  

    main_col = df.columns[0]
    other_cols = df.columns[1:]

    for _, row in df.iterrows():
        main_val = str(row[main_col]).strip()
        if not main_val:
            continue

        for col in other_cols:
            val = str(row[col]).strip()
            if val and val not in ["", "nan", "None"]:
                key = f"{main_val} - {col}"
                result[key] = val

    return result

def extract_email_from_text(text: str) -> str:
    if not text:
        return None

    lines = text.split('\n')

    for line in lines:
        line = line.strip()

        if 'e-mail' in line.lower() or 'email' in line.lower():
            if ':' in line:
                email_part = line.split(':', 1)[1].strip()
                email_clean = email_part.replace(" ", "").strip()

                if '@' in email_clean and len(email_clean) > 3:
                    return email_clean

    text_clean = text.replace(" ", "").replace("\n", " ").strip()

    tokens = re.split(r'[:\s]+', text_clean)
    for token in tokens:
        if '@' in token and '.' in token and len(token) > 3:
            return token

    return None

def merge_employee_email(employee_df: pd.DataFrame, email_df: pd.DataFrame) -> pd.DataFrame:
    result_df = employee_df.copy()

    if email_df.empty:
        return result_df

    email_map = {}
    for _, row in email_df.iterrows():
        page_num = row["page"]
        email = row["E-mail"]
        email_map[page_num] = email

    for page_num, email in email_map.items():
        employee_index = page_num - 1
        if employee_index < len(result_df):
            result_df.loc[employee_index, "E-mail"] = email

    return result_df

def ocr_process(pdf_path: Path, is_employee: bool = False):
    pages = convert_from_path(pdf_path, dpi=200)
    all_text = ""
    page_email_data = []

    for page_num, page in enumerate(pages, start=1):
        text = pytesseract.image_to_string(page, lang=PYTESSERACT_LANG, config=PYTESSERACT_CONFIG)
        all_text += text + "\n\n"

        if is_employee:
            x1, y1, x2, y2 = general_employee_email_coordinate
            cropped = page.crop((x1, y1, x2, y2))
            email_text = pytesseract.image_to_string(cropped, lang="eng", config=PYTESSERACT_CONFIG)

            # Store page number with its email - guaranteed alignment
            page_email_data.append({
                "page": page_num,
                "email_raw": email_text.strip()
            })

    # Clean up pages list after all processing is complete
    del pages
    gc.collect()

    # Return structured email data for employee PDFs
    if is_employee and page_email_data:
        email_df = pd.DataFrame(page_email_data)
        # Extract only email addresses from multi-line text (ignoring phone numbers, etc.)
        email_df["E-mail"] = email_df["email_raw"].apply(extract_email_from_text)
        email_df = email_df[["page", "E-mail"]]
        logger.info(f"Extracted {len(email_df)} email records with guaranteed page alignment")
        return all_text, email_df

    # Return just text for non-employee PDFs or if no emails found (backward compatible)
    if is_employee:
        return all_text, None
    else:
        return all_text

def process_payslip(file_path: Path, file_format: str):
    coordinates = [(94, 293, 431, 457),
        (860, 298, 1611, 545),
        (112, 577, 1609, 1475),
        (113, 1480, 862, 1761),
        (861, 1480, 1614, 1761),
        (111, 1761, 865, 2208),
        (862, 1764, 1613, 2212),
        (112, 1291, 1613, 1472),
        (761, 180, 1652, 229)
    ]

    coordinates_dict = {
        "top_left": coordinates[0],
        "top_right": coordinates[1],
        "middle": coordinates[2],
        "middle_left": coordinates[3],
        "middle_right": coordinates[4],
        "lower_left": coordinates[5],
        "lower_right": coordinates[6],
        "middle_middle": coordinates[7],
        "CVR": coordinates[8]
    }

    pages = convert_from_path(file_path, dpi=200)
    res = []

    for page in pages:
        res_dict = {}

        for section_name, coordinate in coordinates_dict.items():
            x1, y1, x2, y2 = coordinate
            cropped = page.crop((x1, y1, x2, y2))

            if section_name == "top_left":
                text = pytesseract.image_to_string(cropped, lang=PYTESSERACT_LANG, config=PYTESSERACT_CONFIG)
                lines = [line.strip() for line in text.splitlines() if line.strip()]

                name = None
                address_parts = []
                postcode = None

                for line in lines:
                    if not name:
                        name = line
                        continue
                    
                    # Check for postal code (4 digits at start)
                    if re.match(r"^\d{4}", line):
                        postcode = line
                        break  # stop after postcode (no more data below)
                    else:
                        address_parts.append(line)

                # If postcode not found, try the last non-empty line
                if not postcode and address_parts:
                    last_line = address_parts[-1]
                    if re.match(r"^\d{4}", last_line):
                        postcode = last_line
                        address_parts = address_parts[:-1]

                # Fill the result dict
                res_dict["Navn"] = name or ""
                res_dict["Adresse"] = " ".join(address_parts).strip()
                res_dict["Post-nr"] = postcode or ""
                
            if section_name == "top_right":
                text = pytesseract.image_to_string(cropped, lang=PYTESSERACT_LANG, config=PYTESSERACT_CONFIG)
                lines = [line.strip() for line in text.splitlines() if line.strip()]
                for i, line in enumerate(lines):
                    line = line.strip()
                    if ":" in line:
                        key, val = line.split(":", 1)
                        res_dict[key.strip()] = val.strip()
                    else:
                        # handle lines without ':' (continuation lines)
                        if res_dict:
                            last_key = list(res_dict.keys())[-1]
                            res_dict[last_key] += " " + line.strip()

            if section_name == "middle":
                middle_columns_dict = {
                    "Tekst": (2, 747),
                    "Grundlag": (748, 966),
                    "Sats": (965, 1128),
                    "Udbetalt": (1130, 1313),
                    "Trukket": (1313, 1492)
                }

                middle_df = extract_table_from_section(cropped, middle_columns_dict, 20, "dan", False)
                middle_dict = extract_column_from_df(middle_df)
                res_dict.update(middle_dict)

            if section_name == "middle_left":
                middle_left_columns_dict = {
                    "Ferieregnskab": (0, 358),
                    "Perioden": (389, 520),
                    "Ferieår til dato": (531, 748),
                }

                middle_left_df = extract_table_from_section(cropped, middle_left_columns_dict, 20, "dan", False)
                middle_left_dict = extract_column_from_df(middle_left_df)
                res_dict.update(middle_left_dict)

            if section_name == "middle_right":
                middle_right_columns_dict = [
                    {
                        "Ferieregnskab": (2, 601),
                        "Saldo": (651, 751),
                    },
                    {
                        "Ferieregnskab": (2, 346),
                        "Udbetalt": (354, 526),
                        "Trukket": (541, 750),
                    }
                ]

                middle_right_df = extract_table_from_section(cropped, middle_right_columns_dict, 20, "dan", False)
                middle_right_dict = extract_column_from_df(middle_right_df)
                res_dict.update(middle_right_dict)

            if section_name == "lower_left":
                lower_left_columns_dict = {
                    "Saldo": (2, 352),
                    "Perioden": (361, 520),
                    "År til dato": (533, 753),
                }

                lower_left_df = extract_table_from_section(cropped, lower_left_columns_dict, 20, "dan", False)
                lower_left_dict = extract_column_from_df(lower_left_df)
                res_dict.update(lower_left_dict)

            if section_name == "lower_right":
                lower_right_columns_dict = {
                    "Saldo": (2, 352),
                    "Perioden": (361, 520),
                    "År til dato": (533, 753),
                }

                lower_right_df = extract_table_from_section(cropped, lower_right_columns_dict, 20, "dan", False)
                lower_right_dict = extract_column_from_df(lower_right_df)
                res_dict.update(lower_right_dict)
            
            # check if the payslip has additional middle section
            if section_name == "middle_middle":
                middle_middle_columns_dict = {
                    "Ferieafregning, fratrædelse": (0, 401),
                    "Dage": (419, 663),
                    "Brutto": (675, 854),
                    "AM-bidrag": (868, 1058),
                    "A-skat": (1077, 1289),
                    "Netto": (1310, 1499)
                }

                middle_middle_df = extract_table_from_section(cropped, middle_middle_columns_dict, 20, "dan", True)
                if middle_middle_df is not None:
                    middle_middle_dict = extract_column_from_df(middle_middle_df)
                    res_dict.update(middle_middle_dict)
            
            if section_name == "CVR":
                text = pytesseract.image_to_string(cropped, lang="eng", config=PYTESSERACT_CONFIG)
                text = text.strip()

                cvr_match = re.search(r"CVR\s*[:;]?\s*(\d+)", text, re.IGNORECASE)
                if cvr_match:
                    cvr_number = cvr_match.group(1).strip()
                    res_dict["CVR"] = cvr_number
                else:
                    res_dict["CVR"] = ""

        if res_dict and res_dict["Navn"] != "Marianne Trolle" and res_dict["Navn"] != "Lisbeth Hald Anderse":
            res.append(res_dict)
        else:
            print("no record found")

    # Clean up pages list after all processing is complete
    del pages
    gc.collect()

    print("OCR complete")
    # IMPORTANT: Create DataFrame with all columns as strings to preserve leading zeros
    df = pd.DataFrame.from_records(res).astype(str)
    df = df.rename(columns={"Medarbejdernummer": "Medarb.nr."})

    # Clean up NA values in ID columns
    id_columns = ["Medarb.nr.", "CPR-nummer", "CPR", "Post-nr", "CVR"]
    for col in id_columns:
        if col in df.columns:
            df[col] = df[col].replace(['nan', 'None', '<NA>', 'nat'], '')

    return df

def process_employee_general(file_path, file_type):
    # Extract text and email data
    result = ocr_process(file_path, is_employee=True)

    # Handle both tuple and single value return (backward compatibility)
    if isinstance(result, tuple):
        text, email_df = result
    else:
        text = result
        email_df = None

    employees = re.split(r"Frikort brugt:\s*[\d,\.]+", text)
    records = []

    key_value_pattern = re.compile(r"^(.+?)\s+([\w\d,\.%\-]+)$")
    sub_headers = ["Generelt", "Lønoplysninger", "Ansættelsesoplysninger", "Kontakt", "Ferie", "SH", "Fritvalgsordning", "Dage/Timer", "Skattekort"]

    for emp in employees:
        emp = emp.strip()
        if not emp:
            continue

        data = {}
        lines = [line.strip() for line in emp.splitlines() if line.strip()]

        for line in lines:
            # --- Normal colon-based case ---
            if line in sub_headers:
                continue

            if ":" in line:
                key, value = line.split(":", 1)
                key, value = key.strip(), value.strip()

            # --- Case 2: "key value" (no colon) ---
            elif match := key_value_pattern.match(line):
                key, value = match.group(1).strip(), match.group(2).strip()

            else:
                continue

            # Skip if no value or header-like entry
            if not value or key in sub_headers:
                continue

            data[key] = value

        # Add the trailing "Frikort brugt" value if found
        match = re.search(r"Frikort brugt:\s*([\d,\.]+)", emp)
        if match:
            data["Frikort brugt"] = match.group(1)

        if data:
            records.append(data)

    employee_df = pd.DataFrame(records).astype(str)

    if not employee_df.empty:
        employee_df = employee_df.iloc[:, 1:]  # Drop first column if unwanted

    if "Fritvalgsordning 1%" in employee_df.columns:
        employee_df = employee_df.rename(columns={"Fritvalgsordning 1%": "Fritvalgsordning 1 %"})

    # Merge with email data if available
    if email_df is not None and not email_df.empty:
        logger.info(f"Merging {len(email_df)} emails with {len(employee_df)} employees")
        employee_df = merge_employee_email(employee_df, email_df)

    return employee_df

def process_employee_list(file_path, file_type):
    text = ocr_process(file_path)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    
    # Find header and filter lines
    headers = ["Navn", "Medarb.nr.", "Lønperiode", "Adresse", "Adresse 2", "Post-nr. og by"]
    lønperiode_options = {"Måned bagud", "14-dags", "(TOM/FINDES IKKE I DANLØN)"}
    ignore_patterns = ["Din Forsikringsmægler ApS", "Medarbejderliste", "Side "]
    
    # Filter lines
    for i, line in enumerate(lines):
        if line.strip().startswith("Navn"):
            lines = lines[i+1:]
            break
    
    lines = [line for line in lines if not any(pat in line for pat in ignore_patterns)]
    lines = [line.strip().split() for line in lines]
    
    rows = []
    for tokens in lines:
        current = {}

        # --- Navn: all tokens until we hit a number ---
        navn_parts = []
        while tokens and not re.fullmatch(r"\d+", tokens[0]):
            navn_parts.append(tokens.pop(0))
        current["Navn"] = " ".join(navn_parts).strip()

        # --- Medarb.nr. ---
        current["Medarb.nr."] = tokens.pop(0) if tokens and re.fullmatch(r"\d+", tokens[0]) else ""

        # --- Lønperiode: one of the known phrases ---
        lønperiode_parts = []
        while tokens:
            token = tokens.pop(0)
            token = token.strip(".,;:")
            lønperiode_parts.append(token)
            if " ".join(lønperiode_parts) in lønperiode_options:
                break
            
        current["Lønperiode"] = " ".join(lønperiode_parts)

        # --- Post-nr. og by (last two tokens, always starts with digits) ---
        post_idx = None
        for j, tok in enumerate(tokens):
            if re.fullmatch(r"\d{4}", tok):  # exactly 4 digits
                post_idx = j
                break

        if post_idx is not None:
            # Split into adresse + post-nr
            current["Adresse"] = " ".join(tokens[:post_idx]).strip()
            current["Post-nr. og by"] = " ".join(tokens[post_idx:]).strip()
        else:
            # fallback: no post number detected
            current["Adresse"] = " ".join(tokens).strip()
            current["Post-nr. og by"] = ""

        # --- Adresse ---
        current["Adresse"] = " ".join(tokens).strip()
        current["Adresse 2"] = ""  # always blank for now

        rows.append(current)

    # IMPORTANT: Create DataFrame with all columns as strings to preserve leading zeros
    df = pd.DataFrame(rows, columns=headers).astype(str)
    df = df.dropna(subset=["Navn", "Medarb.nr."])
    df = df[df["Navn"].str.strip() != ""]

    return df

def load_danlon_mapping():
    def excel_col_letter(n: int) -> str:
        letters = ""
        while n >= 0:
            n, remainder = divmod(n, 26)
            letters = chr(65 + remainder) + letters
            n -= 1
        return letters

    mapping_path = os.path.join(os.path.dirname(__file__), 'mapping', 'danlon_mapping_new.xlsx')

    try:
        # Read the mapping file
        wb = load_workbook(mapping_path, data_only=False)
        ws = wb["Mapping"]

        rows = list(ws.iter_rows(values_only=False))

        data = [[cell.value for cell in row] for row in rows]

        danlon_mapping_df = pd.DataFrame(data)

        # Extract the mapping rows (based on your original logic)
        mapping_dan_row = danlon_mapping_df.iloc[0, 1:].tolist()
        mapping_eng_row = danlon_mapping_df.iloc[1, 1:].tolist()
        mapping_input_col = danlon_mapping_df.iloc[4, 1:].tolist()
        mapping_file = danlon_mapping_df.iloc[5, 1:].tolist()
        mapping_datatype = danlon_mapping_df.iloc[7, 1:].tolist()
        mapping_guidance = danlon_mapping_df.iloc[8, 1:].tolist()
        mapping_comments = danlon_mapping_df.iloc[10, 1:].tolist()

        # Create excel references
        excel_refs = [excel_col_letter(i) for i in range(len(danlon_mapping_df.columns))]
        excel_refs = excel_refs[1:]

        # Create the final mapping DataFrame
        danlon_mapping = pd.DataFrame({
            "danish": mapping_dan_row,
            "english": mapping_eng_row,
            "input_column": mapping_input_col,
            "file": mapping_file,
            "datatype": mapping_datatype,
            "guidance": mapping_guidance,
            "comments": mapping_comments,
            "excel_ref": excel_refs
        })

        return danlon_mapping

    except Exception as e:
        logger.error(f"Error loading mapping file: {e}")
        logger.error(traceback.format_exc())
        # Return empty mapping if file not found
        return pd.DataFrame()

def extract_years_from_columns(df, pattern):
    """
    Extract all years from column names matching a pattern.
    Returns a sorted list of years (integers).
    """
    years = set()
    for col in df.columns:
        if pattern in str(col):
            # Find 4-digit years like 2024, 2025
            year_match = re.search(r'\b(20\d{2})\b', str(col))
            if year_match:
                years.add(int(year_match.group(1)))
    return sorted(years)

def resolve_holiday_column_names(danlon_mapping, holiday_df):
    """
    Resolve letter guidance (C, D, E, G, H, I) to actual column names from holiday data.
    Updates the input_column field in the mapping dataframe.
    """
    if holiday_df is None or holiday_df.empty:
        logger.info("No holiday data provided, skipping column name resolution")
        return danlon_mapping

    logger.info(f"Resolving holiday column names from guidance letters...")
    logger.info(f"Holiday data columns: {list(holiday_df.columns)}")

    def letter_to_index(letter):
        """Convert Excel column letter to zero-based index (A=0, B=1, C=2, etc.)"""
        if not letter or not isinstance(letter, str):
            return None
        letter = letter.strip().upper()
        if len(letter) == 1 and letter.isalpha():
            return ord(letter) - ord('A')
        return None

    resolved_count = 0

    for idx, row in danlon_mapping.iterrows():
        file_source = str(row.get("file", "")).strip()
        guidance = str(row.get("guidance", "")).strip()

        # Check if this is a holiday data column with letter guidance
        if file_source == "Feriepengeforpligtelse" and guidance:
            col_index = letter_to_index(guidance)

            if col_index is not None and col_index < len(holiday_df.columns):
                actual_column_name = holiday_df.columns[col_index]

                # Update the input_column with the actual column name
                danlon_mapping.at[idx, "input_column"] = actual_column_name

                logger.info(f"Resolved '{row['danish']}': guidance '{guidance}' -> column '{actual_column_name}'")
                resolved_count += 1
            elif col_index is not None:
                logger.warning(f"Column letter '{guidance}' (index {col_index}) exceeds holiday data columns (only {len(holiday_df.columns)} columns)")

    logger.info(f"Resolved {resolved_count} holiday column names from letter guidance")
    return danlon_mapping

def resolve_dynamic_year_columns(danlon_mapping, merged_df, current_year, previous_year):
    """
    Resolve dynamic year patterns like [CURRENT_YEAR] and [PREVIOUS_YEAR] in guidance.
    Updates the input_column field with actual column names from merged data.
    """
    if merged_df is None or merged_df.empty:
        logger.info("No merged data provided, skipping dynamic year column resolution")
        return danlon_mapping

    if current_year is None:
        logger.warning("No current year detected, skipping dynamic year column resolution")
        return danlon_mapping

    logger.info(f"Resolving dynamic year patterns (current={current_year}, previous={previous_year})...")

    resolved_count = 0

    for idx, row in danlon_mapping.iterrows():
        guidance = str(row.get("guidance", "")).strip()

        # Check if guidance contains year placeholders
        if "[CURRENT_YEAR]" in guidance or "[PREVIOUS_YEAR]" in guidance:
            # Replace placeholders with actual years
            resolved_pattern = guidance.replace("[CURRENT_YEAR]", str(current_year))
            if previous_year:
                resolved_pattern = resolved_pattern.replace("[PREVIOUS_YEAR]", str(previous_year))

            # Find matching column in merged data
            matching_col = None
            for col in merged_df.columns:
                if resolved_pattern in str(col):
                    matching_col = col
                    break

            if matching_col:
                # Update the input_column with the actual column name
                danlon_mapping.at[idx, "input_column"] = matching_col
                logger.info(f"Resolved '{row['danish']}': pattern '{guidance}' -> column '{matching_col}'")
                resolved_count += 1
            else:
                logger.warning(f"Could not find column matching pattern '{resolved_pattern}' in merged data")

    logger.info(f"Resolved {resolved_count} dynamic year columns")
    return danlon_mapping

def apply_danlon_mapping(merged_data, danlon_mapping=None, current_year=None, previous_year=None):
    def evaluate_excel_formula(df, formula, danlon_mapping):
        if not isinstance(formula, str):
            return None

        expr = formula.lstrip("=").replace(",", ".")

        tokens = re.findall(r"\b[A-Z]{1,3}\b", expr)

        for token in tokens:
            match = danlon_mapping.loc[danlon_mapping["excel_ref"] == token, "danish"]
            if not match.empty:
                colname = match.iloc[0]
                expr = re.sub(rf"\b{token}\b", f"df['{colname}']", expr)
        try:
            result = eval(expr)
        except Exception as e:
            print(f"Error evaluating formula {formula}: {e}")
            result = None

        return result

    # Load mapping if not provided
    if danlon_mapping is None:
        danlon_mapping = load_danlon_mapping()

    if danlon_mapping.empty:
        print("No mapping file found, returning data as-is")
        return merged_data

    indkomsttype_mapping = {
        "A-Indkomstmodtager/lønansat (kode 00)": "A-INDKOMSTMODTAGER/LØNANSAT (KODE 00)",
        "Anden personlig indkomst/ej lønansat (kode 04)": "ANDEN PERSONLIG INDKOMST/EJ LØNANSAT (kode 04)",
        "B-indkomst modtager/ej lønansat (kode 05)": "B-INDKOMST MODTAGER/EJ LØNANSAT (KODE 05)",
        "§48E (forskerordningen) (kode 08)": "§ 48 E (FORSKERORDNINGEN) (KODE 08)",
        "Skattefri indkomst (kode 09) - udlændinge": "SKATTEFRI INDKOMST (KODE 09)",
        "Skattefri indkomst (kode 09)": "SKATTEFRI INDKOMST (KODE 09)"
    }

    barselsfond_mapping = {
        "Barsel.dk - fuldt medlem": "BARSEL.DK",
        "Frivillig DA barsel": "DA BARSEL",
        "Obligatorisk DA barsel": "DA BARSEL",
        "Barsel.dk - delvis medlem": "BARSEL.DK - DELVIST MEDLEM",
        "Anden barselsordning": "ANDEN BARSELSFOND",
    }

    atp_type_mapping = {
        "Almindelig (A)": "ALMINDELING A",
        "Ingen": None,
    }

    lønperiode_mapping = {
        "Måned bagud": "BAGUDBETALT MÅNEDSLØN",
        "14-dags": "14 DAGE",
    }
    ferieordning_mapping = {
        "Løbende (feriekasse)": "TIMELØNNET (MED FERIEPENGE)",
        "Ingen": "FAST LØN",
        "Ferie med løn": "FAST LØN",
        "Løbende (eget)": "TIMELØNNET (MED FERIEPENGE)"
    }
    
    columns = danlon_mapping["danish"].tolist()

    final_output_df = pd.DataFrame(columns=columns)
    for _, row in danlon_mapping.iterrows():
        try:
            if pd.isna(row["danish"]):
                continue
            print("mapping: ", row["danish"])
            danish_col = row["danish"]
            input_col = row["input_column"]

            if not isinstance(input_col, str) or pd.isna(input_col):
                input_col = ""

            # Build possible column names to look for
            possible_cols = []
            if input_col:
                possible_cols = [input_col, f"{input_col}_df1", f"{input_col}_df2"]

                if "-" in input_col:
                    parts = [p.strip() for p in input_col.split("-")]
                    if len(parts) >= 2:
                        fixed = f"{parts[0]} - {parts[1].capitalize()}"
                        possible_cols.append(fixed)

            # Find matching column from possible column names
            found_col = next((col for col in possible_cols if col in merged_data.columns), None)

            # Special case for timeløn (if not found by exact match)
            if not found_col and input_col and "timeløn" in input_col.lower():
                for col in merged_data.columns:
                    if col and "timeløn" in str(col).lower():
                        found_col = col
                        print(f"Matched '{danish_col}' with column containing 'timeløn': '{col}'")
                        break

            if found_col:
                print(f"found_col raw value: {repr(found_col)}")
                found_col = found_col.strip()           # remove spaces / newlines
                found_col = found_col.replace("\xa0", " ")  # replace non-breaking spaces
                found_col = found_col.replace("-", "-")     # replace non-breaking hyphen
                found_col = found_col.replace("–", "-")     # replace en dash
                found_col = found_col.replace("—", "-")
                print(f"Normalized found_col: '{found_col}'")

                if found_col == "Postnummer og By":
                    merged_data[found_col] = merged_data[found_col].astype(str).str.strip()

                    merged_data["Postnummer_split"] = merged_data[found_col].str.extract(r"(\d+)")
                    merged_data["By_split"] = merged_data[found_col].str.replace(r"\d+", "", regex=True).str.strip()

                    if danish_col == "Postnummer":
                        final_output_df[danish_col] = merged_data["Postnummer_split"]
                        print("Extracted Postnummer")

                    elif danish_col == "By":
                        final_output_df[danish_col] = merged_data["By_split"]
                        print("Extracted By")

                # --- Indkomsttype Mapping ---
                elif found_col == "Indkomsttype":
                    print("found_col is Indkomsttype")
                    merged_data[found_col] = merged_data[found_col].astype(str).str.strip()
                    merged_data[found_col] = (
                        merged_data[found_col]
                        .map(indkomsttype_mapping)
                        .fillna("A-INDKOMSTMODTAGER/LØNANSAT (KODE 00)")
                    )
                    final_output_df[danish_col] = merged_data[found_col]
                    print("Mapping applied for Indkomsttype")

                # --- Barselsfond Mapping ---
                elif found_col == "Barsel":
                    print("found_col is Barselsfond")
                    merged_data[found_col] = merged_data[found_col].astype(str).str.strip()
                    merged_data[found_col] = merged_data[found_col].map(barselsfond_mapping)
                    final_output_df[danish_col] = merged_data[found_col]
                    print("Mapping applied for Barselsfond (unmapped values remain NaN)")

                # --- ATP-type Mapping ---
                elif found_col == "ATP-ordning":
                    print("found_col is ATP-type")
                    merged_data[found_col] = merged_data[found_col].astype(str).str.strip()
                    merged_data[found_col] = (
                        merged_data[found_col]
                        .map(atp_type_mapping)
                        .fillna("ALMINDELING A")
                    )
                    final_output_df[danish_col] = merged_data[found_col]
                    print("Mapping applied for ATP-type")

                # --- Lønperiode Mapping ---
                elif found_col == "Lønperiode":
                    print("found_col is Lønperiode")
                    merged_data[found_col] = merged_data[found_col].astype(str).str.strip()
                    merged_data[found_col] = (
                        merged_data[found_col]
                        .map(lønperiode_mapping)
                        .fillna("FORUDBETAL MÅNEDSLØN")
                    )
                    final_output_df[danish_col] = merged_data[found_col]
                    print("Mapping applied for Lønperiode")
                
                # --- Løntype Mapping ---
                elif found_col == "Ferieordning":
                    print("found_col is Ferieordning")
                    merged_data[found_col] = merged_data[found_col].astype(str).str.strip()
                    merged_data[found_col] = (
                        merged_data[found_col]
                        .map(ferieordning_mapping)
                        .fillna("FAST LØN")
                    )
                    final_output_df[danish_col] = merged_data[found_col]
                    print("Mapping applied for Lønperiode")

                else:
                    final_output_df[danish_col] = merged_data[found_col]

                    # Debug logging for Antal timer
                    if danish_col == "Antal timer":
                        sample_values = final_output_df[danish_col].head(10).tolist()
                        print(f" DEBUG: Mapped 'Antal timer' from column '{found_col}'")
                        print(f" Sample values: {sample_values}")

            else:
                final_output_df[danish_col] = pd.NA

                # Debug logging when column not found
                if danish_col == "Antal timer":
                    print(f"   DEBUG: 'Antal timer' - no matching column found in source data")
                    print(f"   Looking for: {possible_cols}")
                    print(f"   Available columns: {list(merged_data.columns)[:20]}...")  # Show first 20 columns

        except Exception as e:
            print("error occurs: ", e)
    
    if "Ferieordning" in final_output_df.columns:
        final_output_df["Indtægtsart"] = final_output_df["Ferieordning"].apply(
            lambda x: "Direktør" if str(x).strip().lower() == "FAST LØN" else "Medarbejder"
        )
    
    if "Adresse" in final_output_df.columns:
        if "Postnummer" in final_output_df.columns:
            final_output_df["Adresse"] = final_output_df.apply(
                lambda row: str(row["Adresse"]).replace(str(row["Postnummer"]), "")
                if pd.notna(row["Adresse"]) and pd.notna(row["Postnummer"])
                else row["Adresse"],
                axis=1
            )
        if "By" in final_output_df.columns:
            final_output_df["Adresse"] = final_output_df.apply(
                lambda row: str(row["Adresse"]).replace(str(row["By"]), "")
                if pd.notna(row["Adresse"]) and pd.notna(row["By"])
                else row["Adresse"],
                axis=1
            )

    # fill up calculated columns
    calculated_fields = {}

    # Separate formulas into independent and dependent
    formulas_to_process = []
    for _, row in danlon_mapping.iterrows():
        if str(row["input_column"]).strip() == "(calculated field)":
            formula = row.get("datatype")
            is_formula = str(formula).startswith("=")
            if is_formula:
                formulas_to_process.append((row["danish"], formula))

    # Process formulas in dependency order
    max_iterations = len(formulas_to_process) + 1
    iteration = 0

    while formulas_to_process and iteration < max_iterations:
        iteration += 1
        remaining_formulas = []

        for colname, formula in formulas_to_process:
            print(f"\n{colname}")
            print("formula", formula)
            print("colname", colname)

            expr = str(formula).lstrip("=")
            tokens = re.findall(r"\b[A-Z]{1,3}\b", expr)

            # Check if any dependencies are not yet calculated
            has_unresolved_deps = False
            for token in tokens:
                match = danlon_mapping.loc[danlon_mapping["excel_ref"] == token, "danish"]
                if not match.empty:
                    ref_colname = match.iloc[0]
                    # Check if this is a calculated field that hasn't been processed yet
                    if ref_colname in [f[0] for f in formulas_to_process if f[0] != colname]:
                        has_unresolved_deps = True
                        break

            # If has unresolved dependencies, defer to next iteration
            if has_unresolved_deps:
                print(f"  Deferring '{colname}' - has unresolved dependencies")
                remaining_formulas.append((colname, formula))
                continue

            # Track which calculated columns we converted to numeric (need to re-format them)
            converted_calc_fields = []

            # Convert referenced columns to numeric BEFORE evaluating formula
            for token in tokens:
                match = danlon_mapping.loc[danlon_mapping["excel_ref"] == token, "danish"]
                if not match.empty:
                    ref_colname = match.iloc[0]
                    if ref_colname in final_output_df.columns:
                        if ref_colname not in calculated_fields:
                            final_output_df[ref_colname] = final_output_df[ref_colname].apply(parse_european_float_to_numeric)
                            print(f"Converted '{ref_colname}' to numeric for calculation")
                        else:
                            # Parse European format from already-calculated column
                            final_output_df[ref_colname] = final_output_df[ref_colname].apply(parse_european_float_to_numeric)
                            print(f"Re-parsed '{ref_colname}' from European format (already calculated)")
                            converted_calc_fields.append(ref_colname)

            res = evaluate_excel_formula(final_output_df, formula, danlon_mapping)
            final_output_df[colname] = res.apply(format_european_float) if hasattr(res, 'apply') else format_european_float(res)
            # Clean up: remove thousand separators and fix double commas
            final_output_df[colname] = final_output_df[colname].apply(clean_float_string)
            calculated_fields[colname] = True

            sample_values = final_output_df[colname].head(3).tolist()
            print(f"  Calculated '{colname}' sample values: {sample_values}")

            for converted_col in converted_calc_fields:
                final_output_df[converted_col] = final_output_df[converted_col].apply(format_european_float)
                # Clean up: remove thousand separators and fix double commas
                final_output_df[converted_col] = final_output_df[converted_col].apply(clean_float_string)
                print(f"  Re-formatted '{converted_col}' back to European format after use")

        formulas_to_process = remaining_formulas

    # post processing after calculating calculated columns
    for calc_col in calculated_fields.keys():
        if calc_col in final_output_df.columns:
            # Check current dtype
            current_dtype = final_output_df[calc_col].dtype
            print(f"'{calc_col}': current dtype = {current_dtype}")

            # If it's numeric, re-format to European format
            if current_dtype in ['float64', 'int64', 'float32', 'int32']:
                final_output_df[calc_col] = final_output_df[calc_col].apply(format_european_float)
                # Clean up: remove thousand separators and fix double commas
                final_output_df[calc_col] = final_output_df[calc_col].apply(clean_float_string)
                print(f"Re-formatted '{calc_col}' from {current_dtype} to European string format")

            # Log sample values
            sample = final_output_df[calc_col].head(3).tolist()
            print(f"    Sample values: {sample}")

    for _, row in danlon_mapping.iterrows():
        try:
            danish_col = row["danish"]

            if pd.isna(danish_col) or danish_col not in final_output_df.columns:
                continue

            datatype_raw = row.get("datatype", "")

            if danish_col in calculated_fields:
                continue

            datatype_str = str(datatype_raw).strip().lower()

            if datatype_str == "float":
                final_output_df[danish_col] = final_output_df[danish_col].apply(format_european_float)
                # Clean up: remove thousand separators and fix double commas
                final_output_df[danish_col] = final_output_df[danish_col].apply(clean_float_string)

            elif datatype_str == "string":
                final_output_df[danish_col] = final_output_df[danish_col].replace([pd.NA, np.nan, None], '').astype(str).replace(['nan', 'None', '<NA>'], '')

                if danish_col in ["Ansættelsesdato", "Anciennitetsdato"]:
                    final_output_df[danish_col] = final_output_df[danish_col].astype(str).str.replace('.', '-', regex=False)

                # Kontonummer should always be 10 digits with leading zeros
                if danish_col == "Kontonummer":
                    final_output_df[danish_col] = final_output_df[danish_col].apply(
                        lambda x: str(x).strip().zfill(10) if str(x).strip() and str(x).strip().isdigit() else x
                    )
                    print(f"Formatted 'Kontonummer' to 10 digits with leading zeros")

        except Exception as e:
            print(f"Error formatting column '{danish_col}': {e}")
            continue

    # Final cleanup: Remove all NaN values and replace with empty strings
    print("\n🧹 Final cleanup - removing NaN values...")

    # Step 1: Replace actual NaN/NA values with empty strings
    final_output_df = final_output_df.fillna('')

    # Step 2: Replace string representations of NaN with empty strings
    final_output_df = final_output_df.replace([pd.NA, np.nan, None, '<NA>', 'nan', 'NaN', 'None', 'nat', 'NaT'], '')

    # Step 3: For each column, clean up any remaining "nan" strings
    for col in final_output_df.columns:
        if col in final_output_df.columns:
            # Replace string "nan" values
            final_output_df[col] = final_output_df[col].apply(
                lambda x: '' if str(x).lower() in ['nan', 'none', '<na>', 'nat'] else x
            )

    print("Final cleanup complete")

    # Filter by "Er aktiv" column - only keep rows with "Ja"
    if "Er aktiv" in final_output_df.columns:
        initial_rows = len(final_output_df)

        # Keep only rows where "Er aktiv" equals "Ja"
        final_output_df = final_output_df[
            final_output_df["Er aktiv"].astype(str).str.strip().str.upper() == "JA"
        ]

        filtered_rows = len(final_output_df)
        removed_count = initial_rows - filtered_rows

        if removed_count > 0:
            print(f"Filtered out {removed_count} inactive employees (Er aktiv != 'Ja') ({initial_rows} → {filtered_rows} rows)")
        else:
            print(f"All {initial_rows} employees are active")

    # Remove duplicate rows based on Medarbejdernummer 1
    if "Medarbejdernummer 1" in final_output_df.columns:
        initial_rows = len(final_output_df)

        # Remove duplicates, keeping the first occurrence
        final_output_df = final_output_df.drop_duplicates(subset=["Medarbejdernummer 1"], keep="first")

        final_rows = len(final_output_df)
        removed_duplicates = initial_rows - final_rows

        if removed_duplicates > 0:
            print(f"Removed {removed_duplicates} duplicate rows based on Medarbejdernummer 1 ({initial_rows} → {final_rows} rows)")
        else:
            print(f"No duplicate employee numbers found")

    # Special case: Calculate CY and DF columns using Ferietillæg from payslip
    # CY = CX - "Ferietillæg til gode for [current_year] - Saldo"
    # DF = DE - "Ferietillæg til gode for [current_year] - Saldo"
    if current_year is not None and previous_year is not None:
        print(f"\n🔧 Calculating special columns CY and DF using current year {current_year}...")

        curr_ferietillag_col = None
        prev_ferietillag_col = None
        
        for col in merged_data.columns:
            if f"Ferietillæg til gode for {current_year}" in str(col) and "Saldo" in str(col):
                curr_ferietillag_col = col
                
            if f"Ferietillæg til gode for {previous_year}" in str(col) and "Saldo" in str(col):
                prev_ferietillag_col = col
                
            if prev_ferietillag_col is not None and curr_ferietillag_col is not None:
                break

        if curr_ferietillag_col:
            print(f"   Found Ferietillæg column: '{curr_ferietillag_col}'")

            cx_col = "Indeværende ferieår, Optjent ferietillæg"
            cy_col = "Indeværende ferieår, brugt ferietillæg"

            if cx_col in final_output_df.columns and cy_col in final_output_df.columns:
                try:
                    # Convert to numeric for calculation
                    cx_values = final_output_df[cx_col].apply(parse_european_float_to_numeric)
                    ferietillag_values = merged_data[curr_ferietillag_col].apply(parse_european_float_to_numeric)

                    # Calculate CY = CX - Ferietillæg (clamp to 0 if negative)
                    result = cx_values - ferietillag_values
                    result = result.apply(lambda x: max(0, x) if pd.notna(x) else x)

                    # Format back to European format
                    final_output_df[cy_col] = result.apply(format_european_float).apply(clean_float_string)

                    print(f"   Calculated '{cy_col}' = '{cx_col}' - '{curr_ferietillag_col}'")
                    sample = final_output_df[cy_col].head(3).tolist()
                    print(f"      Sample values: {sample}")
                except Exception as e:
                    print(f"   Error calculating CY column: {e}")
        else:
            print(f"   Could not find 'Ferietillæg til gode for {current_year} - Saldo' column in merged data")
            print(f"   Available Ferietillæg columns: {[c for c in merged_data.columns if 'Ferietillæg' in str(c)]}")
        
        if prev_ferietillag_col:

            de_col = "Sidste ferieår, Optjent ferietillæg"
            df_col = "Sidste ferieår, Brugt ferietillæg"

            if de_col in final_output_df.columns and df_col in final_output_df.columns:
                try:
                    # Convert to numeric for calculation
                    de_values = final_output_df[de_col].apply(parse_european_float_to_numeric)
                    ferietillag_values = merged_data[prev_ferietillag_col].apply(parse_european_float_to_numeric)

                    # Calculate DF = DE - Ferietillæg (clamp to 0 if negative)
                    result = de_values - ferietillag_values
                    result = result.apply(lambda x: max(0, x) if pd.notna(x) else x)

                    # Format back to European format
                    final_output_df[df_col] = result.apply(format_european_float).apply(clean_float_string)

                    print(f"   Calculated '{df_col}' = '{de_col}' - '{prev_ferietillag_col}'")
                    sample = final_output_df[df_col].head(3).tolist()
                    print(f"      Sample values: {sample}")
                except Exception as e:
                    print(f" Error calculating DF column: {e}")
        else:
            print(f"   Could not find 'Ferietillæg til gode for {previous_year} - Saldo' column in merged data")
            print(f"   Available Ferietillæg columns: {[c for c in merged_data.columns if 'Ferietillæg' in str(c)]}")

    return final_output_df

def merge(left: pd.DataFrame, right: pd.DataFrame, on: str, how="left"):
    merged = pd.merge(left, right, on=on, how=how, suffixes=("_left", "_right"))
    overlap_cols = set(left.columns).intersection(right.columns) - {on}

    for col in overlap_cols:
        left_col = f"{col}_left"
        right_col = f"{col}_right"

        merged[col] = merged[left_col].combine_first(merged[right_col])
        merged.drop(columns=[left_col, right_col], inplace=True)

    return merged

def merge_and_map_data(processed_data):
    # Load mapping and resolve holiday column names BEFORE merging
    danlon_mapping = load_danlon_mapping()

    if danlon_mapping.empty:
        logger.error("Failed to load mapping file")
        return None

    # Extract years and validate consistency between holiday and payslip data
    current_year = None
    previous_year = None

    if "employee_holiday" in processed_data and "employee_payslip" in processed_data:
        holiday_df = processed_data["employee_holiday"]
        payslip_df = processed_data["employee_payslip"]

        # Extract years from both datasets
        holiday_years = extract_years_from_columns(holiday_df, "Optjent dage")
        payslip_years = extract_years_from_columns(payslip_df, "Ferietillæg til gode for")

        logger.info(f"Detected years - Holiday: {holiday_years}, Payslip: {payslip_years}")

        # Validate years match
        if holiday_years and payslip_years:
            if set(holiday_years) != set(payslip_years):
                error_msg = (
                    f"YEAR MISMATCH DETECTED!\n"
                    f"Holiday data contains years: {holiday_years}\n"
                    f"Payslip data contains years: {payslip_years}\n"
                    f"Please verify you have uploaded the correct files for the same time period."
                )
                logger.error(error_msg)
                raise ValueError(error_msg)

            # Set current and previous years
            sorted_years = sorted(holiday_years, reverse=True)
            current_year = sorted_years[0] if len(sorted_years) > 0 else None
            previous_year = sorted_years[1] if len(sorted_years) > 1 else None

            logger.info(f"Year validation passed - Current: {current_year}, Previous: {previous_year}")
        else:
            logger.warning("Could not extract years from data files")

    # Resolve holiday column names from letter guidance
    if "employee_holiday" in processed_data:
        holiday_df = processed_data["employee_holiday"]
        danlon_mapping = resolve_holiday_column_names(danlon_mapping, holiday_df)

    id_columns = ["Medarb.nr.", "CPR-nummer", "CPR", "Post-nr", "Postnummer", "CVR"]

    for df_name, df in processed_data.items():
        if df is not None and not df.empty:
            for col in id_columns:
                if col in df.columns:
                    df[col] = df[col].astype(str)
                    logging.info(f"Converted '{col}' to string in {df_name}")
                    if col == "CPR-nummer":
                        logging.info(f"CPR nummer row {df[col].head()}")

    key = None

    # Start with the main employee data if available
    if "employee_active" in processed_data:
        key = "employee_active"
        merged = processed_data["employee_active"].copy()
    elif "employee_general" in processed_data:
        key = "employee_general"
        merged = processed_data["employee_general"].copy()
    elif "employee_payslip" in processed_data:
        # If no employee data, create empty DataFrame with payslip data
        key = "employee_payslip"
        merged = processed_data.get("employee_payslip", pd.DataFrame())
    else:
        return None

    # Merge with holiday data if available
    if "employee_holiday" in processed_data and not merged.empty:
        logging.info("merging employee holiday")
        holiday_df = processed_data["employee_holiday"]
        if "Navn" in merged.columns and "Navn" in holiday_df.columns:
            merged = merge(merged, holiday_df, on="Navn", how="left")
    
    # Merge with employee list if available
    if "employee_list" in processed_data and not merged.empty:
        logging.info("merging employee list")
        employee_list_df = processed_data["employee_list"]
        if "Medarb.nr." in merged.columns and "Medarb.nr." in employee_list_df.columns:
            # Ensure consistent data types
            merged["Medarb.nr."] = merged["Medarb.nr."].astype(str)
            employee_list_df["Medarb.nr."] = employee_list_df["Medarb.nr."].astype(str)
            merged = merge(merged, employee_list_df, on="Medarb.nr.", how="left")
    
    # Merge with employee data if available
    if "employee_general" in processed_data and not merged.empty and key != "employee_general":
        logging.info("merging employee general")
        employee_df = processed_data["employee_general"]
        if "CPR-nummer" in employee_df.columns:
            sample_cpr = employee_df['CPR-nummer'].dropna().head(5).tolist()
            logging.info(f"CPR nummer before merge with employee_general: {sample_cpr}")
        if "Medarb.nr." in merged.columns and "Medarb.nr." in employee_df.columns:
            merged["Medarb.nr."] = merged["Medarb.nr."].astype(str)
            employee_df["Medarb.nr."] = employee_df["Medarb.nr."].astype(str)
            merged = merge(merged, employee_df, on="Medarb.nr.", how="left")
    
    if "employee_payslip" in processed_data and not merged.empty and key != "employee_payslip":
        logging.info("merging employee_payslip")
        payslip_df = processed_data["employee_payslip"]
        if "Medarb.nr." in merged.columns and "Medarb.nr." in payslip_df.columns:
            merged["Medarb.nr."] = merged["Medarb.nr."].astype(str)
            payslip_df["Medarb.nr."] = payslip_df["Medarb.nr."].astype(str)
            merged = merge(merged, payslip_df, on="Medarb.nr.", how="left")

    if merged.empty:
        largest_key = max(processed_data.keys(), key=lambda k: len(processed_data[k]), default=None)
        if largest_key:
            merged = processed_data[largest_key]

    if not merged.empty:
        for col in merged.columns:
            if col and 'CPR' in str(col).upper():
                sample_cpr = merged[col].dropna().head(10).astype(str).tolist()
                logging.info(f"After merging, before mapping - Column '{col}' dtype={merged[col].dtype}, values: {sample_cpr}")

    # Resolve dynamic year columns after merging is complete
    if not merged.empty and current_year is not None:
        logging.info("Resolving dynamic year column references...")
        danlon_mapping = resolve_dynamic_year_columns(danlon_mapping, merged, current_year, previous_year)

    if not merged.empty:
        logging.info("start mapping merged employee")
        mapped_result = apply_danlon_mapping(merged, danlon_mapping, current_year, previous_year)

    return mapped_result